import json
import logging
import os

import click
import queue
import re
import sqlite3
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Optional

import fitz
from flask import (
    Flask,
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    stream_with_context,
    url_for,
)

from api_clients.ozon_client import OzonApiError
from api_clients.wb_client import WbApiError
from api_clients.wb_content_client import WbContentError
from config import Config
from database.catalog_db import init_catalog_db
from database.db import get_db_connection, init_db
from services.orders_service import OrdersService
from services.wb_catalog_service import run_full_wb_catalog_sync
from utils.helpers import parse_int
from utils.ozon_label_cache import (
    load_cached_ozon_label,
    normalize_ozon_label_pdf,
    save_cached_ozon_label,
)


def setup_logging():
    os.makedirs(os.path.dirname(Config.LOG_PATH), exist_ok=True)
    logging.basicConfig(
        filename=Config.LOG_PATH,
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )


def create_app():
    setup_logging()
    init_db()
    init_catalog_db()

    app = Flask(__name__)
    app.secret_key = Config.FLASK_SECRET_KEY

    @app.context_processor
    def inject_static_version():
        css = os.path.join(Config.BASE_DIR, "static", "style.css")
        js = os.path.join(Config.BASE_DIR, "static", "app.js")
        stamps = []
        for p in (css, js):
            try:
                stamps.append(int(os.path.getmtime(p)))
            except OSError:
                pass
        return {"static_v": max(stamps) if stamps else int(time.time())}

    @app.after_request
    def _static_revalidate_js_css(resp):
        path = request.path or ""
        if path.startswith("/static/") and path.endswith((".js", ".css")):
            resp.headers["Cache-Control"] = "no-cache"
        return resp

    service = OrdersService()

    def _sync_active_orders_with_retry() -> None:
        """
        Ozon может отдавать старый статус сразу после ship-вызова.
        Делаем несколько коротких попыток синхронизации 2 активных статусов.
        """
        last_exc: Exception | None = None
        for attempt in range(3):
            try:
                service.sync_from_ozon(
                    statuses=["awaiting_packaging", "awaiting_deliver"],
                    since=None,
                    to=None,
                    limit=100,
                    max_records=5000,
                )
                return
            except Exception as exc:
                last_exc = exc
                logging.exception(
                    "Авто-синхронизация после ship не удалась (попытка %s/3)",
                    attempt + 1,
                )
                if attempt < 2:
                    time.sleep(1.2)
        if last_exc:
            raise last_exc

    def _normalize_pick_article(raw_offer_id: str) -> str:
        """
        Примеры:
        BASE_3_1_697_B_M  -> 3_1_697_B
        BASE_3_1_697_BB_M -> 3_1_697_B
        """
        s = str(raw_offer_id or "").strip()
        if not s:
            return "—"
        def _latinize_color(c: str) -> str:
            t = str(c or "").upper()[:1]
            return (
                t.replace("А", "A")
                .replace("В", "B")
                .replace("Е", "E")
                .replace("К", "K")
                .replace("М", "M")
                .replace("Н", "H")
                .replace("О", "O")
                .replace("Р", "P")
                .replace("С", "C")
                .replace("Т", "T")
                .replace("У", "Y")
                .replace("Х", "X")
            )

        m = re.search(r"^[^_]+_(\d+)_(\d+)_(\d+)_([A-Za-zА-Яа-я])", s)
        if m:
            return f"{m.group(1)}_{m.group(2)}_{m.group(3)}_{_latinize_color(m.group(4))}"
        parts = s.split("_")
        if len(parts) >= 5:
            tail = _latinize_color(parts[4] if parts[4] else "X")
            return f"{parts[1]}_{parts[2]}_{parts[3]}_{tail}"
        return s

    @app.get("/")
    def index():
        summary = service.get_summary()
        return render_template("index.html", summary=summary)

    @app.get("/logo.png")
    def logo_png():
        return send_file(os.path.join(app.root_path, "logo.png"), mimetype="image/png")

    @app.get("/icon-ozon.png")
    def icon_ozon_png():
        return send_file(os.path.join(app.root_path, "icon-ozon.png"), mimetype="image/png")

    @app.get("/logo-wb.png")
    def logo_wb_png():
        return send_file(os.path.join(app.root_path, "logowb.png"), mimetype="image/png")

    @app.get("/orders")
    @app.get("/orders_ozon")
    def orders():
        status = request.args.get("status", "all")
        date_from = request.args.get("date_from", "")
        date_to = request.args.get("date_to", "")
        query = request.args.get("q", "")

        per_page = Config.ORDERS_CHUNK_SIZE
        data = service.get_orders(
            status=status,
            date_from=date_from or None,
            date_to=date_to or None,
            query=query or None,
            page=1,
            per_page=per_page,
            marketplace="ozon",
        )
        top_shipments = service.get_shipments_with_awaiting_deliver_orders()
        has_more = data["page"] < data["pages"]
        return render_template(
            "orders.html",
            data=data,
            top_shipments=top_shipments,
            filters={
                "status": status,
                "date_from": date_from,
                "date_to": date_to,
                "q": query,
            },
            has_more=has_more,
            next_page=2 if has_more else None,
        )

    @app.get("/orders/load-more")
    def orders_load_more():
        status = request.args.get("status", "all")
        date_from = request.args.get("date_from", "")
        date_to = request.args.get("date_to", "")
        query = request.args.get("q", "")
        page = parse_int(request.args.get("page"), 2, min_value=2)
        per_page = Config.ORDERS_CHUNK_SIZE

        data = service.get_orders(
            status=status,
            date_from=date_from or None,
            date_to=date_to or None,
            query=query or None,
            page=page,
            per_page=per_page,
            marketplace="ozon",
        )
        html = render_template("partials/order_rows.html", orders=data["orders"])
        has_more = page < data["pages"]
        next_page = page + 1 if has_more else None
        return jsonify({"html": html, "has_more": has_more, "next_page": next_page})

    @app.get("/orders_wb")
    def orders_wb():
        status = request.args.get("status", "new")
        date_from = request.args.get("date_from", "")
        date_to = request.args.get("date_to", "")
        query = request.args.get("q", "")
        wh = request.args.get("wh", "")
        per_page = Config.ORDERS_CHUNK_SIZE
        data = service.get_orders(
            status=status,
            date_from=date_from or None,
            date_to=date_to or None,
            query=query or None,
            page=1,
            per_page=per_page,
            marketplace="wb",
            warehouse=wh or None,
        )
        top_shipments = service.get_wb_supplies_with_orders()
        has_more = data["page"] < data["pages"]
        return render_template(
            "orders_wb.html",
            data=data,
            top_shipments=top_shipments,
            filters={
                "status": status,
                "date_from": date_from,
                "date_to": date_to,
                "q": query,
                "wh": wh,
            },
            has_more=has_more,
            next_page=2 if has_more else None,
        )

    @app.get("/orders_wb/load-more")
    def orders_wb_load_more():
        status = request.args.get("status", "new")
        date_from = request.args.get("date_from", "")
        date_to = request.args.get("date_to", "")
        query = request.args.get("q", "")
        wh = request.args.get("wh", "")
        page = parse_int(request.args.get("page"), 2, min_value=2)
        per_page = Config.ORDERS_CHUNK_SIZE
        data = service.get_orders(
            status=status,
            date_from=date_from or None,
            date_to=date_to or None,
            query=query or None,
            page=page,
            per_page=per_page,
            marketplace="wb",
            warehouse=wh or None,
        )
        html = render_template("partials/order_rows.html", orders=data["orders"])
        has_more = page < data["pages"]
        next_page = page + 1 if has_more else None
        return jsonify({"html": html, "has_more": has_more, "next_page": next_page})

    @app.post("/orders_wb/update")
    def update_orders_wb():
        since = request.form.get("since", "")
        to = request.form.get("to", "")
        try:
            result = service.sync_from_wb(since=since or None, to=to or None, max_records=5000)
            flash(
                f"Заказы WB обновлены. Создано: {result['created']}, обновлено: {result['updated']}, "
                f"удалено устаревших: {result.get('deleted', 0)}.",
                "success",
            )
        except WbApiError as exc:
            logging.exception("Ошибка WB API")
            flash(str(exc), "error")
        except sqlite3.OperationalError as exc:
            logging.exception("Ошибка SQLite при обновлении заказов WB")
            if "locked" in str(exc).lower():
                flash(
                    "База данных занята (возможно, уже идёт обновление). Подождите и повторите.",
                    "error",
                )
            else:
                flash(str(exc), "error")
        except Exception as exc:
            logging.exception("Неожиданная ошибка при обновлении заказов WB")
            flash(
                (str(exc) or "").strip()
                or "Не удалось выполнить синхронизацию Wildberries (см. logs/app.log).",
                "error",
            )
        return redirect(url_for("orders_wb", status="new"))

    @app.post("/orders_wb/sync-catalog-new")
    def sync_wb_catalog_new_orders():
        """Content API: название/фото/размер/категория только для заказов вкладки «Новые»."""
        try:
            stats = service.sync_wb_catalog_for_new_orders_only()
            if stats.get("message"):
                flash(stats["message"], "info")
            else:
                flash(
                    f"Каталог для «Новые»: уникальных nm {stats.get('unique_nm', 0)}, "
                    f"карточек {stats.get('cards_fetched', 0)}, "
                    f"обновлено позиций {stats.get('items_updated', 0)}.",
                    "success" if stats.get("ok") else "warning",
                )
            if stats.get("error"):
                flash(str(stats["error"]), "warning")
        except Exception:
            logging.exception("sync_wb_catalog_for_new_orders_only")
            flash("Не удалось подтянуть карточки для заказов «Новые»", "error")
        return redirect(url_for("orders_wb", status="new"))

    @app.post("/orders_wb/sync-catalog-new-stream")
    def sync_wb_catalog_new_orders_stream():
        """NDJSON: прогресс «Подтянуть карточки для Новые» (Content API + обновление order_items)."""

        def generate():
            q: queue.Queue = queue.Queue()

            def worker():
                try:
                    def cb(ev):
                        q.put(("progress", ev))

                    result = service.sync_wb_catalog_for_new_orders_only(progress_cb=cb)
                    q.put(("done", result))
                except WbContentError as exc:
                    q.put(("err", str(exc)))
                except sqlite3.OperationalError as exc:
                    if "locked" in str(exc).lower():
                        q.put(
                            (
                                "err",
                                "База данных занята. Подождите и повторите.",
                            )
                        )
                    else:
                        q.put(("err", str(exc)))
                except Exception:
                    logging.exception("Стрим: sync_wb_catalog_for_new_orders_only")
                    q.put(("err", "Не удалось подтянуть карточки"))

            thread = threading.Thread(target=worker, daemon=True)
            thread.start()
            while True:
                kind, payload = q.get()
                if kind == "progress":
                    line = json.dumps({"type": "progress", **payload}, ensure_ascii=False) + "\n"
                    yield line
                elif kind == "done":
                    yield json.dumps({"type": "done", "result": payload}, ensure_ascii=False) + "\n"
                    break
                elif kind == "err":
                    yield json.dumps({"type": "error", "message": payload}, ensure_ascii=False) + "\n"
                    break
            thread.join(timeout=2.0)

        return Response(
            stream_with_context(generate()),
            mimetype="application/x-ndjson",
            headers={
                "Cache-Control": "no-cache",
                "X-Accel-Buffering": "no",
            },
        )

    @app.post("/orders_wb/sync-full-catalog-stream")
    def sync_wb_full_catalog_stream():
        """NDJSON: полная синхронизация каталога WB в wb_catalog.db (Content API, постранично)."""

        def generate():
            cancel_event = threading.Event()
            q: queue.Queue = queue.Queue()

            def worker():
                try:

                    def cb(ev):
                        q.put(("progress", ev))

                    result = run_full_wb_catalog_sync(
                        progress_cb=cb,
                        cancel_check=cancel_event.is_set,
                    )
                    try:
                        if not result.get("skipped"):
                            q.put(
                                (
                                    "progress",
                                    {"step": "wb_full_catalog_enrich_start"},
                                )
                            )
                            enrich_stats = service.enrich_wb_order_items_from_local_catalog()
                            result["order_items_enriched"] = int(
                                enrich_stats.get("items_updated") or 0
                            )
                            q.put(
                                (
                                    "progress",
                                    {
                                        "step": "wb_full_catalog_enrich_done",
                                        "items_updated": int(result["order_items_enriched"]),
                                    },
                                )
                            )
                    except Exception:
                        logging.exception(
                            "После синхронизации каталога WB: enrich_wb_order_items_from_local_catalog"
                        )
                    q.put(("done", result))
                except sqlite3.OperationalError as exc:
                    if "locked" in str(exc).lower():
                        q.put(
                            (
                                "err",
                                "База данных занята. Подождите и повторите.",
                            )
                        )
                    else:
                        q.put(("err", str(exc)))
                except Exception:
                    logging.exception("Стрим: run_full_wb_catalog_sync")
                    q.put(("err", "Не удалось синхронизировать каталог WB"))

            thread = threading.Thread(target=worker, daemon=True)
            thread.start()
            try:
                while True:
                    kind, payload = q.get()
                    if kind == "progress":
                        line = json.dumps({"type": "progress", **payload}, ensure_ascii=False) + "\n"
                        yield line
                    elif kind == "done":
                        yield json.dumps({"type": "done", "result": payload}, ensure_ascii=False) + "\n"
                        break
                    elif kind == "err":
                        yield json.dumps({"type": "error", "message": payload}, ensure_ascii=False) + "\n"
                        break
            except GeneratorExit:
                cancel_event.set()
                raise
            finally:
                cancel_event.set()
                thread.join(timeout=90.0)

        return Response(
            stream_with_context(generate()),
            mimetype="application/x-ndjson",
            headers={
                "Cache-Control": "no-cache",
                "X-Accel-Buffering": "no",
            },
        )

    @app.post("/orders_wb/update-json")
    def update_orders_wb_json():
        since = request.form.get("since", "")
        to = request.form.get("to", "")
        try:
            result = service.sync_from_wb(since=since or None, to=to or None, max_records=5000)
            return jsonify(
                {
                    "ok": True,
                    "message": (
                        f"Готово. Создано: {result.get('created', 0)}, "
                        f"обновлено: {result.get('updated', 0)}, "
                        f"удалено: {result.get('deleted', 0)}."
                    ),
                    "result": result,
                }
            )
        except WbApiError as exc:
            logging.exception("Ошибка WB API")
            return jsonify({"ok": False, "message": str(exc)}), 400
        except sqlite3.OperationalError as exc:
            logging.exception("Ошибка SQLite при обновлении заказов WB")
            if "locked" in str(exc).lower():
                msg = (
                    "База данных занята (возможно, уже идёт обновление). "
                    "Подождите и не запускайте синхронизацию дважды."
                )
            else:
                msg = str(exc)
            return jsonify({"ok": False, "message": msg}), 503
        except Exception as exc:
            logging.exception("Неожиданная ошибка при обновлении заказов WB")
            detail = (str(exc) or "").strip()
            return (
                jsonify(
                    {
                        "ok": False,
                        "message": detail
                        if detail
                        else "Не удалось выполнить синхронизацию Wildberries (см. logs/app.log).",
                    }
                ),
                500,
            )

    @app.post("/orders_wb/update-stream")
    def update_orders_wb_stream():
        """NDJSON-стрим: строки {type: progress|done|error, ...} для модалки прогресса."""
        since = request.form.get("since", "")
        to = request.form.get("to", "")

        def generate():
            q: queue.Queue = queue.Queue()

            def worker():
                try:
                    def cb(ev):
                        q.put(("progress", ev))

                    result = service.sync_from_wb(
                        since=since or None,
                        to=to or None,
                        max_records=5000,
                        progress_cb=cb,
                    )
                    q.put(("done", result))
                except WbApiError as exc:
                    q.put(("wb_err", str(exc)))
                except sqlite3.OperationalError as exc:
                    if "locked" in str(exc).lower():
                        q.put(
                            (
                                "err",
                                "База данных занята (возможно, уже идёт обновление). "
                                "Подождите и не запускайте синхронизацию дважды.",
                            )
                        )
                    else:
                        q.put(("err", str(exc)))
                except Exception as exc:
                    logging.exception("Стрим: ошибка синхронизации WB")
                    detail = (str(exc) or "").strip()
                    q.put(
                        (
                            "err",
                            detail
                            if detail
                            else "Не удалось выполнить синхронизацию Wildberries (см. logs/app.log).",
                        )
                    )

            thread = threading.Thread(target=worker, daemon=True)
            thread.start()
            while True:
                kind, payload = q.get()
                if kind == "progress":
                    line = json.dumps({"type": "progress", **payload}, ensure_ascii=False) + "\n"
                    yield line
                elif kind == "done":
                    yield json.dumps({"type": "done", "result": payload}, ensure_ascii=False) + "\n"
                    break
                elif kind == "wb_err":
                    yield json.dumps({"type": "error", "message": payload}, ensure_ascii=False) + "\n"
                    break
                elif kind == "err":
                    yield json.dumps({"type": "error", "message": payload}, ensure_ascii=False) + "\n"
                    break
            thread.join(timeout=2.0)

        return Response(
            stream_with_context(generate()),
            mimetype="application/x-ndjson",
            headers={
                "Cache-Control": "no-cache",
                "X-Accel-Buffering": "no",
            },
        )

    @app.post("/orders/update")
    def update_orders():
        status = request.form.get("status", "all")
        sync_scope = (request.form.get("sync_scope") or "active").strip()
        since = request.form.get("since", "")
        to = request.form.get("to", "")
        limit = parse_int(request.form.get("limit"), 100, min_value=1, max_value=200)
        sync_statuses = (
            ["awaiting_packaging", "awaiting_deliver", "delivering"]
            if sync_scope == "all"
            else ["awaiting_packaging", "awaiting_deliver"]
        )

        try:
            result = service.sync_from_ozon(
                status=status,
                statuses=sync_statuses,
                since=since or None,
                to=to or None,
                limit=limit,
                max_records=5000,
            )
            flash(
                f"Данные успешно обновлены. Создано: {result['created']}, обновлено: {result['updated']}.",
                "success",
            )
        except OzonApiError as exc:
            logging.exception("Ошибка Ozon API")
            flash(str(exc), "error")
        except sqlite3.OperationalError as exc:
            logging.exception("Ошибка SQLite при обновлении заказов")
            if "locked" in str(exc).lower():
                flash(
                    "База данных занята (возможно, уже идёт обновление). Подождите и повторите.",
                    "error",
                )
            else:
                flash(str(exc), "error")
        except Exception:
            logging.exception("Неожиданная ошибка при обновлении заказов")
            flash("Не удалось получить данные из Ozon API", "error")

        return redirect(url_for("orders"))

    @app.post("/orders/update-json")
    def update_orders_json():
        status = request.form.get("status", "all")
        sync_scope = (request.form.get("sync_scope") or "active").strip()
        since = request.form.get("since", "")
        to = request.form.get("to", "")
        limit = parse_int(request.form.get("limit"), 100, min_value=1, max_value=200)
        sync_statuses = (
            ["awaiting_packaging", "awaiting_deliver", "delivering"]
            if sync_scope == "all"
            else ["awaiting_packaging", "awaiting_deliver"]
        )

        try:
            result = service.sync_from_ozon(
                status=status,
                statuses=sync_statuses,
                since=since or None,
                to=to or None,
                limit=limit,
                max_records=5000,
            )
            return jsonify(
                {
                    "ok": True,
                    "scope": sync_scope,
                    "message": (
                        f"Готово. Создано: {result.get('created', 0)}, "
                        f"обновлено: {result.get('updated', 0)}, "
                        f"удалено: {result.get('deleted', 0)}."
                    ),
                    "result": result,
                }
            )
        except OzonApiError as exc:
            logging.exception("Ошибка Ozon API")
            return jsonify({"ok": False, "message": str(exc)}), 400
        except sqlite3.OperationalError as exc:
            logging.exception("Ошибка SQLite при обновлении заказов")
            if "locked" in str(exc).lower():
                msg = (
                    "База данных занята (возможно, уже идёт обновление). "
                    "Подождите и не запускайте синхронизацию дважды."
                )
            else:
                msg = str(exc)
            return jsonify({"ok": False, "message": msg}), 503
        except Exception:
            logging.exception("Неожиданная ошибка при обновлении заказов")
            return jsonify({"ok": False, "message": "Не удалось получить данные из Ozon API"}), 500

    @app.post("/orders/update-stream")
    def update_orders_stream():
        status = request.form.get("status", "all")
        sync_scope = (request.form.get("sync_scope") or "active").strip()
        since = request.form.get("since", "")
        to = request.form.get("to", "")
        limit = parse_int(request.form.get("limit"), 100, min_value=1, max_value=200)
        sync_statuses = (
            ["awaiting_packaging", "awaiting_deliver", "delivering"]
            if sync_scope == "all"
            else ["awaiting_packaging", "awaiting_deliver"]
        )

        def generate():
            q: queue.Queue = queue.Queue()

            def worker():
                try:
                    result = service.sync_from_ozon(
                        status=status,
                        statuses=sync_statuses,
                        since=since or None,
                        to=to or None,
                        limit=limit,
                        max_records=5000,
                        progress_cb=lambda ev: q.put(("progress", ev)),
                    )
                    q.put(("done", result))
                except OzonApiError as exc:
                    q.put(("err", str(exc)))
                except sqlite3.OperationalError as exc:
                    if "locked" in str(exc).lower():
                        q.put(
                            (
                                "err",
                                "База данных занята (возможно, уже идёт обновление). Подождите и повторите.",
                            )
                        )
                    else:
                        q.put(("err", str(exc)))
                except Exception:
                    logging.exception("Стрим: update_orders_stream")
                    q.put(("err", "Не удалось получить данные из Ozon API"))

            thread = threading.Thread(target=worker, daemon=True)
            thread.start()
            while True:
                kind, payload = q.get()
                if kind == "progress":
                    yield json.dumps({"type": "progress", **payload}, ensure_ascii=False) + "\n"
                elif kind == "done":
                    yield json.dumps({"type": "done", "result": payload}, ensure_ascii=False) + "\n"
                    break
                else:
                    yield json.dumps({"type": "error", "message": payload}, ensure_ascii=False) + "\n"
                    break
            thread.join(timeout=2.0)

        return Response(
            stream_with_context(generate()),
            mimetype="application/x-ndjson",
            headers={
                "Cache-Control": "no-cache",
                "X-Accel-Buffering": "no",
            },
        )

    @app.get("/orders/<int:order_id>/label.pdf")
    def order_label_pdf(order_id: int):
        rel = service.ensure_order_label_pdf_file(order_id)
        if not rel:
            flash("Нет этикетки: у позиций заказа нет штрихкода.", "error")
            mp = service.get_order_marketplace(order_id)
            fallback = url_for("orders_wb", status="new") if mp == "wb" else url_for("orders")
            return redirect(request.referrer or fallback)
        path = os.path.join(Config.BASE_DIR, rel.replace("/", os.sep))
        posting = service.get_order_posting_number(order_id) or str(order_id)
        safe = re.sub(r"[^\w\-.]+", "_", posting, flags=re.UNICODE)[:120]
        dl = f"etiketka_{safe}.pdf"
        return send_file(path, mimetype="application/pdf", as_attachment=True, download_name=dl)

    @app.get("/orders/<int:order_id>/ozon-label.pdf")
    def order_ozon_label_pdf(order_id: int):
        info = service.get_order_posting_and_status(order_id)
        if not info or not info.get("posting_number"):
            flash("Заказ не найден или нет posting_number.", "error")
            return redirect(request.referrer or url_for("orders"))
        if info.get("status") != "awaiting_deliver":
            flash("Этикетка Ozon доступна только для статуса «Ожидает отгрузки».", "error")
            return redirect(request.referrer or url_for("orders"))

        posting = info["posting_number"]
        try:
            pdf = service.client.get_fbs_package_label_pdf(posting)
        except OzonApiError as exc:
            logging.exception("Ошибка получения этикетки Ozon")
            flash(str(exc), "error")
            return redirect(request.referrer or url_for("orders"))

        pdf = normalize_ozon_label_pdf(pdf)
        save_cached_ozon_label(posting, pdf)

        safe = re.sub(r"[^\w\-.]+", "_", posting, flags=re.UNICODE)[:120]
        dl = f"ozon_posting_label_{safe}.pdf"
        from io import BytesIO

        return send_file(
            BytesIO(pdf),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=dl,
        )

    @app.get("/orders/<int:order_id>/wb-label.png")
    def order_wb_label_png(order_id: int):
        conn = get_db_connection()
        try:
            row = conn.execute(
                "SELECT marketplace, posting_number, wb_label_path FROM orders WHERE id = ?",
                (order_id,),
            ).fetchone()
        finally:
            conn.close()
        if not row or str(row["marketplace"] or "") != "wb":
            flash("WB заказ не найден.", "error")
            return redirect(url_for("orders_wb", status="confirm"))
        rel = (row["wb_label_path"] or "").strip()
        if not rel:
            flash("Этикетка WB для заказа пока недоступна.", "error")
            return redirect(url_for("orders_wb", status="confirm"))
        path = os.path.join(Config.BASE_DIR, rel.replace("/", os.sep))
        if not os.path.isfile(path):
            flash("Файл WB-этикетки не найден.", "error")
            return redirect(url_for("orders_wb", status="confirm"))
        posting = str(row["posting_number"] or order_id)
        return send_file(
            path,
            mimetype="image/png",
            as_attachment=True,
            download_name=f"wb_order_label_{posting}.png",
        )

    @app.get("/shipments/<int:shipment_id>/orders-tape.pdf")
    def shipment_orders_tape_pdf(shipment_id: int):
        data = service.get_shipment_detail(shipment_id)
        shipment = data.get("shipment") or {}
        orders = data.get("orders") or []
        if not shipment or not orders:
            flash("Поставка не найдена или в ней нет заказов.", "error")
            return redirect(url_for("orders"))

        out_doc = fitz.open()
        try:
            prepared: List[dict] = []
            issues: List[str] = []

            # Preflight: verify local labels and prepare Ozon label retrieval list.
            missing_ozon: List[str] = []
            for order in orders:
                order_id = int(order.get("id") or 0)
                posting = str(order.get("posting_number") or "")
                if not order_id or not posting:
                    issues.append(f"Некорректный заказ в поставке (id={order.get('id')}).")
                    continue

                rel = service.ensure_order_label_pdf_file(order_id)
                if not rel:
                    issues.append(f"{posting}: нет локальной этикетки (ШК).")
                    continue
                local_path = os.path.join(Config.BASE_DIR, rel.replace("/", os.sep))
                if not os.path.isfile(local_path):
                    issues.append(f"{posting}: файл локальной этикетки не найден.")
                    continue

                ozon_pdf = load_cached_ozon_label(posting)
                if not ozon_pdf:
                    missing_ozon.append(posting)

                prepared.append(
                    {
                        "posting": posting,
                        "local_path": local_path,
                        "ozon_pdf": ozon_pdf,
                    }
                )

            # Load missing Ozon labels in parallel (faster on large shipments).
            if missing_ozon:
                unique_missing = list(dict.fromkeys(missing_ozon))
                worker_count = min(8, max(2, len(unique_missing)))

                def fetch_one(posting_number: str) -> tuple[str, bytes]:
                    pdf = service.client.get_fbs_package_label_pdf(posting_number)
                    pdf = normalize_ozon_label_pdf(pdf)
                    save_cached_ozon_label(posting_number, pdf)
                    return posting_number, pdf

                fetched: dict[str, bytes] = {}
                failed_postings: List[str] = []
                with ThreadPoolExecutor(max_workers=worker_count) as pool:
                    fut_map = {pool.submit(fetch_one, pn): pn for pn in unique_missing}
                    for fut in as_completed(fut_map):
                        pn = fut_map[fut]
                        try:
                            got_pn, pdf = fut.result()
                            fetched[got_pn] = pdf
                        except Exception:
                            failed_postings.append(pn)

                # Этикетка недоступна: отменённый в Ozon posting удаляем локально и
                # собираем ленту без него; настоящая ошибка по-прежнему блокирует сборку.
                cancelled_excluded: List[str] = []
                for pn in failed_postings:
                    if service.client.get_fbs_posting_status(pn) == "cancelled":
                        service.delete_order_by_posting(pn)
                        cancelled_excluded.append(pn)
                    else:
                        issues.append(f"{pn}: не удалось получить этикетку Ozon.")

                if cancelled_excluded:
                    excluded_set = set(cancelled_excluded)
                    prepared = [
                        e for e in prepared if str(e.get("posting") or "") not in excluded_set
                    ]
                    flash(
                        "Заказы отменены Ozon и исключены из ленты: "
                        + ", ".join(cancelled_excluded),
                        "warning",
                    )

                for entry in prepared:
                    if entry.get("ozon_pdf"):
                        continue
                    pn = str(entry.get("posting") or "")
                    if pn in fetched:
                        entry["ozon_pdf"] = fetched[pn]

            if issues:
                preview = "; ".join(issues[:5])
                tail = f" (и еще {len(issues) - 5})" if len(issues) > 5 else ""
                flash(
                    "Лента заказов не сформирована: не все этикетки доступны. "
                    f"Проблемы: {preview}{tail}",
                    "error",
                )
                return redirect(url_for("shipment_detail", shipment_id=shipment_id))

            # Build final tape only after all labels are pre-validated.
            for entry in prepared:
                ozon_doc = fitz.open(stream=entry["ozon_pdf"], filetype="pdf")
                # Make local label page visually comparable to Ozon label size in tape PDF.
                target_rect = ozon_doc[0].rect if ozon_doc.page_count else fitz.Rect(0, 0, 580, 400)

                local_doc = fitz.open(entry["local_path"])
                for i in range(local_doc.page_count):
                    dst = out_doc.new_page(width=target_rect.width, height=target_rect.height)
                    dst.show_pdf_page(dst.rect, local_doc, i, keep_proportion=True)
                local_doc.close()

                out_doc.insert_pdf(ozon_doc)
                ozon_doc.close()

            if out_doc.page_count == 0:
                flash("Не удалось собрать ленту заказов: нет подходящих этикеток.", "error")
                return redirect(url_for("shipment_detail", shipment_id=shipment_id))

            from io import BytesIO

            # Final compression for interleaved tape: fonts/images/object streams.
            out_doc.subset_fonts()
            pdf_bytes = out_doc.tobytes(
                garbage=4,
                deflate=True,
                deflate_images=True,
                deflate_fonts=True,
                use_objstms=1,
            )
            safe_name = re.sub(r"[^\w\-.]+", "_", str(shipment.get("name") or shipment_id), flags=re.UNICODE)[:120]
            filename = f"lenta_zakazov_{safe_name}.pdf"
            return send_file(
                BytesIO(pdf_bytes),
                mimetype="application/pdf",
                as_attachment=True,
                download_name=filename,
            )
        except OzonApiError as exc:
            logging.exception("Ошибка Ozon API при сборке ленты заказов")
            flash(str(exc), "error")
            return redirect(url_for("shipment_detail", shipment_id=shipment_id))
        except Exception as exc:
            logging.exception("Ошибка при сборке ленты заказов")
            flash(str(exc) or "Не удалось сформировать ленту заказов.", "error")
            return redirect(url_for("shipment_detail", shipment_id=shipment_id))
        finally:
            out_doc.close()

    @app.get("/shipments/<int:shipment_id>/picklist.xlsx")
    def shipment_picklist_xlsx(shipment_id: int):
        data = service.get_shipment_detail(shipment_id)
        shipment = data.get("shipment") or {}
        orders = data.get("orders") or []
        if not shipment or not orders:
            flash("Поставка не найдена или в ней нет заказов.", "error")
            return redirect(url_for("orders"))

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from io import BytesIO

            agg: dict[str, int] = {}
            for order in orders:
                for item in (order.get("items") or []):
                    offer_id = str(item.get("offer_id") or item.get("sku") or "").strip()
                    norm = _normalize_pick_article(offer_id)
                    qty = int(item.get("quantity") or 0)
                    if qty <= 0:
                        qty = 1
                    agg[norm] = agg.get(norm, 0) + qty

            wb = Workbook()
            ws_print = wb.active
            ws_print.title = "Печать"
            ws_print.append(["Артикул", "Количество"])
            total_qty = 0
            for article, qty in sorted(agg.items(), key=lambda x: x[0]):
                ws_print.append([article, qty])
                total_qty += int(qty or 0)
            ws_print.append(["Итого", total_qty])

            ws_print.column_dimensions["A"].width = 28
            ws_print.column_dimensions["B"].width = 14

            thin = Side(style="thin", color="000000")
            all_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            base_font = Font(name="Calibri", size=18)
            bold_font = Font(name="Calibri", size=18, bold=True)

            max_row = ws_print.max_row
            for row in ws_print.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=2):
                for c in row:
                    c.font = base_font
                    c.border = all_border
                    c.alignment = Alignment(vertical="center")

            for c in ws_print[1]:
                c.font = bold_font
                c.alignment = Alignment(horizontal="center", vertical="center")

            for r in range(2, max_row + 1):
                ws_print.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
                ws_print.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")

            ws_print.cell(row=max_row, column=1).font = bold_font
            ws_print.cell(row=max_row, column=2).font = bold_font

            # Sheet 2: "Сборка" (pivot-like by brand-prefix -> color-letter -> size)
            ws_pick = wb.create_sheet("Сборка")
            shipment_title = f"Поставка: {shipment.get('name') or shipment_id}"
            ws_pick.append([shipment_title, ""])
            ws_pick.append(["Названия строк", "Количество"])
            ws_pick.column_dimensions["A"].width = 34
            ws_pick.column_dimensions["B"].width = 16

            def _parse_pick_keys(offer_id: str, size_raw: str) -> tuple[str, str, str]:
                s = str(offer_id or "").strip()
                prefix = (s.split("_", 1)[0] if "_" in s else s) or "—"
                parts = s.split("_")
                color = "X"
                size = str(size_raw or "").strip()

                def _latinize_color(c: str) -> str:
                    t = str(c or "").upper()[:1]
                    return (
                        t.replace("А", "A")
                        .replace("В", "B")
                        .replace("Е", "E")
                        .replace("К", "K")
                        .replace("М", "M")
                        .replace("Н", "H")
                        .replace("О", "O")
                        .replace("Р", "P")
                        .replace("С", "C")
                        .replace("Т", "T")
                        .replace("У", "Y")
                        .replace("Х", "X")
                    )

                # Expected forms:
                # BASE_6_2_104_B_S
                # L_3_1_44_BS
                # L_3_1_7_WM
                # L_3_1_58_ВM
                # BASE_3_1_697_BB_M
                if len(parts) >= 5:
                    tail = parts[4] or ""
                    color = _latinize_color(tail[:1] if tail else "X")
                    if len(parts) >= 6 and parts[5]:
                        size = parts[5]
                    elif len(tail) > 1:
                        size = tail[1:]

                if not size:
                    size = parts[-1] if parts else "—"
                return prefix, color, size.upper()

            tree: dict[str, dict[str, dict[str, int]]] = {}
            for order in orders:
                for item in (order.get("items") or []):
                    offer_id = str(item.get("offer_id") or item.get("sku") or "").strip()
                    qty = int(item.get("quantity") or 0)
                    if qty <= 0:
                        qty = 1
                    prefix, color, size = _parse_pick_keys(offer_id, str(item.get("manufacturer_size") or ""))
                    tree.setdefault(prefix, {}).setdefault(color, {})
                    tree[prefix][color][size] = tree[prefix][color].get(size, 0) + qty

            ws_pick["A1"].font = bold_font
            ws_pick["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws_pick.merge_cells("A1:B1")
            ws_pick["A2"].font = bold_font
            ws_pick["B2"].font = bold_font
            ws_pick["A2"].alignment = Alignment(horizontal="left", vertical="center")
            ws_pick["B2"].alignment = Alignment(horizontal="center", vertical="center")
            fill_header = PatternFill(fill_type="solid", fgColor="D9E1F2")
            fill_level_1 = PatternFill(fill_type="solid", fgColor="E2F0D9")
            fill_level_2 = PatternFill(fill_type="solid", fgColor="FCE4D6")
            for c in ws_pick[2]:
                c.fill = fill_header

            def _size_sort_key(sz: str) -> tuple[int, str]:
                order_map = {
                    "XS": 1,
                    "S": 2,
                    "M": 3,
                    "L": 4,
                    "XL": 5,
                    "XXL": 6,
                    "XXXL": 7,
                    "4XL": 8,
                    "5XL": 9,
                    "6XL": 10,
                }
                return (order_map.get(sz.upper(), 100), sz)

            grand_total = 0
            for prefix in sorted(tree.keys()):
                prefix_total = sum(
                    qty
                    for c in tree[prefix].values()
                    for qty in c.values()
                )
                grand_total += prefix_total
                ws_pick.append([prefix, prefix_total])
                pr = ws_pick.max_row
                ws_pick[f"A{pr}"].font = bold_font
                ws_pick[f"B{pr}"].font = bold_font
                ws_pick[f"A{pr}"].fill = fill_level_1
                ws_pick[f"B{pr}"].fill = fill_level_1

                for color in sorted(tree[prefix].keys()):
                    color_total = sum(tree[prefix][color].values())
                    ws_pick.append([f"  Цвет-{color}", color_total])
                    cr = ws_pick.max_row
                    ws_pick[f"A{cr}"].font = bold_font
                    ws_pick[f"B{cr}"].font = bold_font
                    ws_pick[f"A{cr}"].fill = fill_level_2
                    ws_pick[f"B{cr}"].fill = fill_level_2

                    for size in sorted(tree[prefix][color].keys(), key=_size_sort_key):
                        ws_pick.append([f"    {size}", tree[prefix][color][size]])

            ws_pick.append(["Итого", grand_total])
            tr = ws_pick.max_row
            ws_pick[f"A{tr}"].font = bold_font
            ws_pick[f"B{tr}"].font = bold_font
            ws_pick[f"A{tr}"].fill = fill_header
            ws_pick[f"B{tr}"].fill = fill_header

            for row in ws_pick.iter_rows(min_row=2, max_row=ws_pick.max_row, min_col=1, max_col=2):
                for c in row:
                    if c.font != bold_font:
                        c.font = base_font
                    c.border = all_border
                    c.alignment = Alignment(vertical="center")
            for r in range(2, ws_pick.max_row + 1):
                ws_pick.cell(row=r, column=2).alignment = Alignment(horizontal="right", vertical="center")

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            safe_name = re.sub(r"[^\w\-.]+", "_", str(shipment.get("name") or shipment_id), flags=re.UNICODE)[:120]
            filename = f"fayl_podbora_{safe_name}.xlsx"
            return send_file(
                bio,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )
        except Exception:
            logging.exception("Не удалось сформировать файл подбора")
            flash("Не удалось сформировать файл подбора.", "error")
            return redirect(url_for("shipment_detail", shipment_id=shipment_id))

    @app.get("/shipments/wb/<string:supply_id>/orders-tape.pdf")
    def wb_shipment_orders_tape_pdf(supply_id: str):
        data = service.get_wb_shipment_detail_by_supply_id(supply_id)
        shipment = data.get("shipment") or {}
        orders = data.get("orders") or []
        if not shipment or not orders:
            flash("Поставка WB не найдена или в ней нет заказов.", "error")
            return redirect(url_for("orders_wb", status="confirm"))

        out_doc = fitz.open()
        try:
            issues: List[str] = []
            prepared: List[dict] = []
            for order in orders:
                order_id = int(order.get("id") or 0)
                posting = str(order.get("posting_number") or order_id or "")
                if not order_id:
                    issues.append(f"Некорректный заказ в поставке (id={order.get('id')}).")
                    continue
                rel = service.ensure_order_label_pdf_file(order_id)
                if not rel:
                    issues.append(f"{posting}: нет локальной этикетки (ШК).")
                    continue
                local_path = os.path.join(Config.BASE_DIR, rel.replace("/", os.sep))
                if not os.path.isfile(local_path):
                    issues.append(f"{posting}: файл локальной этикетки не найден.")
                    continue
                wb_rel = str(order.get("wb_label_path") or "").strip()
                if not wb_rel:
                    issues.append(f"{posting}: нет WB-этикетки заказа.")
                    continue
                wb_path = os.path.join(Config.BASE_DIR, wb_rel.replace("/", os.sep))
                if not os.path.isfile(wb_path):
                    issues.append(f"{posting}: файл WB-этикетки не найден.")
                    continue
                prepared.append({"posting": posting, "local_path": local_path, "wb_path": wb_path})

            if issues:
                preview = "; ".join(issues[:5])
                tail = f" (и еще {len(issues) - 5})" if len(issues) > 5 else ""
                flash(
                    "Лента заказов WB не сформирована: не все этикетки доступны. "
                    f"Проблемы: {preview}{tail}",
                    "error",
                )
                return redirect(url_for("wb_shipment_detail", supply_id=supply_id))

            for entry in prepared:
                wb_doc_raw = fitz.open(entry["wb_path"])
                wb_doc = wb_doc_raw
                if not wb_doc_raw.is_pdf:
                    wb_doc = fitz.open("pdf", wb_doc_raw.convert_to_pdf())
                target_rect = wb_doc[0].rect if wb_doc.page_count else fitz.Rect(0, 0, 580, 400)
                local_doc = fitz.open(entry["local_path"])
                for i in range(local_doc.page_count):
                    dst = out_doc.new_page(width=target_rect.width, height=target_rect.height)
                    dst.show_pdf_page(dst.rect, local_doc, i, keep_proportion=True)
                local_doc.close()
                out_doc.insert_pdf(wb_doc)
                wb_doc.close()
                wb_doc_raw.close()

            if out_doc.page_count == 0:
                flash("Не удалось собрать ленту заказов WB: нет подходящих этикеток.", "error")
                return redirect(url_for("wb_shipment_detail", supply_id=supply_id))

            from io import BytesIO

            out_doc.subset_fonts()
            pdf_bytes = out_doc.tobytes(
                garbage=4,
                deflate=True,
                deflate_images=True,
                deflate_fonts=True,
                use_objstms=1,
            )
            safe_name = re.sub(r"[^\w\-.]+", "_", str(shipment.get("name") or supply_id), flags=re.UNICODE)[:120]
            filename = f"lenta_zakazov_{safe_name}.pdf"
            return send_file(
                BytesIO(pdf_bytes),
                mimetype="application/pdf",
                as_attachment=True,
                download_name=filename,
            )
        except Exception as exc:
            logging.exception("Ошибка при сборке ленты заказов WB")
            flash(str(exc) or "Не удалось сформировать ленту заказов WB.", "error")
            return redirect(url_for("wb_shipment_detail", supply_id=supply_id))
        finally:
            out_doc.close()

    @app.get("/shipments/wb/<string:supply_id>/picklist.xlsx")
    def wb_shipment_picklist_xlsx(supply_id: str):
        data = service.get_wb_shipment_detail_by_supply_id(supply_id)
        shipment = data.get("shipment") or {}
        orders = data.get("orders") or []
        if not shipment or not orders:
            flash("Поставка WB не найдена или в ней нет заказов.", "error")
            return redirect(url_for("orders_wb", status="confirm"))

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from io import BytesIO

            agg: dict[str, int] = {}
            for order in orders:
                for item in (order.get("items") or []):
                    offer_id = str(item.get("offer_id") or item.get("sku") or "").strip()
                    norm = _normalize_pick_article(offer_id)
                    qty = int(item.get("quantity") or 0)
                    if qty <= 0:
                        qty = 1
                    agg[norm] = agg.get(norm, 0) + qty

            wb = Workbook()
            ws_print = wb.active
            ws_print.title = "Печать"
            ws_print.append(["Артикул", "Количество"])
            total_qty = 0
            for article, qty in sorted(agg.items(), key=lambda x: x[0]):
                ws_print.append([article, qty])
                total_qty += int(qty or 0)
            ws_print.append(["Итого", total_qty])

            ws_print.column_dimensions["A"].width = 28
            ws_print.column_dimensions["B"].width = 14

            thin = Side(style="thin", color="000000")
            all_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            base_font = Font(name="Calibri", size=18)
            bold_font = Font(name="Calibri", size=18, bold=True)

            max_row = ws_print.max_row
            for row in ws_print.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=2):
                for c in row:
                    c.font = base_font
                    c.border = all_border
                    c.alignment = Alignment(vertical="center")

            for c in ws_print[1]:
                c.font = bold_font
                c.alignment = Alignment(horizontal="center", vertical="center")

            for r in range(2, max_row + 1):
                ws_print.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
                ws_print.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")

            ws_print.cell(row=max_row, column=1).font = bold_font
            ws_print.cell(row=max_row, column=2).font = bold_font

            ws_pick = wb.create_sheet("Сборка")
            shipment_title = f"Поставка: {shipment.get('name') or supply_id}"
            ws_pick.append([shipment_title, ""])
            ws_pick.append(["Названия строк", "Количество"])
            ws_pick.column_dimensions["A"].width = 34
            ws_pick.column_dimensions["B"].width = 16

            def _parse_pick_keys(offer_id: str, size_raw: str) -> tuple[str, str, str]:
                s = str(offer_id or "").strip()
                prefix = (s.split("_", 1)[0] if "_" in s else s) or "—"
                parts = s.split("_")
                color = "X"
                size = str(size_raw or "").strip()

                def _latinize_color(c: str) -> str:
                    t = str(c or "").upper()[:1]
                    return (
                        t.replace("А", "A")
                        .replace("В", "B")
                        .replace("Е", "E")
                        .replace("К", "K")
                        .replace("М", "M")
                        .replace("Н", "H")
                        .replace("О", "O")
                        .replace("Р", "P")
                        .replace("С", "C")
                        .replace("Т", "T")
                        .replace("У", "Y")
                        .replace("Х", "X")
                    )

                if len(parts) >= 5:
                    tail = parts[4] or ""
                    color = _latinize_color(tail[:1] if tail else "X")
                    if len(parts) >= 6 and parts[5]:
                        size = parts[5]
                    elif len(tail) > 1:
                        size = tail[1:]

                if not size:
                    size = parts[-1] if parts else "—"
                return prefix, color, size.upper()

            tree: dict[str, dict[str, dict[str, int]]] = {}
            for order in orders:
                for item in (order.get("items") or []):
                    offer_id = str(item.get("offer_id") or item.get("sku") or "").strip()
                    qty = int(item.get("quantity") or 0)
                    if qty <= 0:
                        qty = 1
                    prefix, color, size = _parse_pick_keys(offer_id, str(item.get("manufacturer_size") or ""))
                    tree.setdefault(prefix, {}).setdefault(color, {})
                    tree[prefix][color][size] = tree[prefix][color].get(size, 0) + qty

            ws_pick["A1"].font = bold_font
            ws_pick["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws_pick.merge_cells("A1:B1")
            ws_pick["A2"].font = bold_font
            ws_pick["B2"].font = bold_font
            ws_pick["A2"].alignment = Alignment(horizontal="left", vertical="center")
            ws_pick["B2"].alignment = Alignment(horizontal="center", vertical="center")
            fill_header = PatternFill(fill_type="solid", fgColor="D9E1F2")
            fill_level_1 = PatternFill(fill_type="solid", fgColor="E2F0D9")
            fill_level_2 = PatternFill(fill_type="solid", fgColor="FCE4D6")
            for c in ws_pick[2]:
                c.fill = fill_header

            def _size_sort_key(sz: str) -> tuple[int, str]:
                order_map = {
                    "XS": 1,
                    "S": 2,
                    "M": 3,
                    "L": 4,
                    "XL": 5,
                    "XXL": 6,
                    "XXXL": 7,
                    "4XL": 8,
                    "5XL": 9,
                    "6XL": 10,
                }
                return (order_map.get(sz.upper(), 100), sz)

            grand_total = 0
            for prefix in sorted(tree.keys()):
                prefix_total = sum(qty for c in tree[prefix].values() for qty in c.values())
                grand_total += prefix_total
                ws_pick.append([prefix, prefix_total])
                pr = ws_pick.max_row
                ws_pick[f"A{pr}"].font = bold_font
                ws_pick[f"B{pr}"].font = bold_font
                ws_pick[f"A{pr}"].fill = fill_level_1
                ws_pick[f"B{pr}"].fill = fill_level_1

                for color in sorted(tree[prefix].keys()):
                    color_total = sum(tree[prefix][color].values())
                    ws_pick.append([f"  Цвет-{color}", color_total])
                    cr = ws_pick.max_row
                    ws_pick[f"A{cr}"].font = bold_font
                    ws_pick[f"B{cr}"].font = bold_font
                    ws_pick[f"A{cr}"].fill = fill_level_2
                    ws_pick[f"B{cr}"].fill = fill_level_2

                    for size in sorted(tree[prefix][color].keys(), key=_size_sort_key):
                        ws_pick.append([f"    {size}", tree[prefix][color][size]])

            ws_pick.append(["Итого", grand_total])
            tr = ws_pick.max_row
            ws_pick[f"A{tr}"].font = bold_font
            ws_pick[f"B{tr}"].font = bold_font
            ws_pick[f"A{tr}"].fill = fill_header
            ws_pick[f"B{tr}"].fill = fill_header

            for row in ws_pick.iter_rows(min_row=2, max_row=ws_pick.max_row, min_col=1, max_col=2):
                for c in row:
                    if c.font != bold_font:
                        c.font = base_font
                    c.border = all_border
                    c.alignment = Alignment(vertical="center")
            for r in range(2, ws_pick.max_row + 1):
                ws_pick.cell(row=r, column=2).alignment = Alignment(horizontal="right", vertical="center")

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            safe_name = re.sub(r"[^\w\-.]+", "_", str(shipment.get("name") or supply_id), flags=re.UNICODE)[:120]
            filename = f"fayl_podbora_{safe_name}.xlsx"
            return send_file(
                bio,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )
        except Exception:
            logging.exception("Не удалось сформировать файл подбора WB")
            flash("Не удалось сформировать файл подбора.", "error")
            return redirect(url_for("wb_shipment_detail", supply_id=supply_id))

    @app.get("/orders/shipment/next-name")
    def orders_shipment_next_name():
        name = service.suggest_next_shipment_name(marketplace="ozon")
        return jsonify({"name": name})

    @app.get("/orders/shipment/available")
    def orders_shipment_available():
        shipments = service.get_shipments_available_for_awaiting_deliver()
        return jsonify({"shipments": shipments})

    @app.post("/orders/shipment/create")
    def orders_shipment_create():
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        ids_raw = request.form.getlist("order_ids")
        next_url = request.form.get("next") or url_for("orders")
        if isinstance(next_url, str) and next_url.startswith("/") and not next_url.startswith("//"):
            pass
        else:
            next_url = url_for("orders")
        shipment_name = (request.form.get("shipment_name") or "").strip()
        order_ids = []
        for x in ids_raw:
            try:
                order_ids.append(int(x))
            except (TypeError, ValueError):
                continue
        try:
            result = service.create_shipment_with_orders(
                shipment_name,
                order_ids,
                ship_after=True,
            )
            if result.get("ok"):
                # Always refresh active Ozon statuses right after create/ship:
                # awaiting_packaging + awaiting_deliver (independent of split/non-split flow).
                try:
                    _sync_active_orders_with_retry()
                    service.attach_split_children_to_shipment(
                        shipment_id=int(result.get("shipment_id") or 0),
                        source_postings=result.get("source_postings") or [],
                    )
                except Exception:
                    logging.exception("Не удалось выполнить авто-синхронизацию после отгрузки")
                    msg = "Поставка создана, но авто-обновление заказов не удалось. Нажмите «Обновить заказы»."
                    if is_ajax:
                        return jsonify({"ok": False, "message": msg, "next_url": next_url}), 500
                    flash(msg, "error")
                failed = result.get("failed_postings") or []
                success_msg = (
                    f"Поставка «{result['name']}»: отгружено {result.get('count', 0)} "
                    f"из {result.get('requested_count', result.get('count', 0))}."
                )
                if failed:
                    success_msg += f" Пропущено из-за ошибок Ozon: {len(failed)}."
                if is_ajax:
                    return jsonify(
                        {
                            "ok": True,
                            "message": success_msg,
                            "next_url": next_url,
                            "failed_count": len(failed),
                            "failed_postings": failed[:10],
                        }
                    )
                flash(success_msg, "success")
            else:
                err = result.get("error") or "Не удалось создать поставку"
                if is_ajax:
                    return jsonify({"ok": False, "message": err, "next_url": next_url}), 400
                flash(err, "error")
        except OzonApiError as exc:
            logging.exception("Ошибка Ozon API при создании поставки/отгрузки")
            if is_ajax:
                return jsonify({"ok": False, "message": str(exc), "next_url": next_url}), 500
            flash(str(exc), "error")
        return redirect(next_url)

    @app.post("/orders/shipment/add-existing")
    def orders_shipment_add_existing():
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        ids_raw = request.form.getlist("order_ids")
        next_url = request.form.get("next") or url_for("orders")
        if isinstance(next_url, str) and next_url.startswith("/") and not next_url.startswith("//"):
            pass
        else:
            next_url = url_for("orders")

        shipment_id_raw = request.form.get("shipment_id")
        try:
            shipment_id = int(shipment_id_raw)
        except (TypeError, ValueError):
            if is_ajax:
                return jsonify({"ok": False, "message": "Не выбрана поставка", "next_url": next_url}), 400
            flash("Не выбрана поставка", "error")
            return redirect(next_url)

        order_ids: List[int] = []
        for x in ids_raw:
            try:
                order_ids.append(int(x))
            except (TypeError, ValueError):
                continue

        try:
            result = service.add_orders_to_existing_shipment(
                shipment_id=shipment_id,
                order_ids=order_ids,
            )
            if result.get("ok"):
                # Always refresh active Ozon statuses right after add/ship:
                # awaiting_packaging + awaiting_deliver (independent of split/non-split flow).
                try:
                    _sync_active_orders_with_retry()
                    service.attach_split_children_to_shipment(
                        shipment_id=int(result.get("shipment_id") or shipment_id),
                        source_postings=result.get("source_postings") or [],
                    )
                except Exception:
                    logging.exception("Не удалось выполнить авто-синхронизацию после добавления в поставку")
                    msg = "Заказы добавлены в поставку, но авто-обновление не удалось. Нажмите «Обновить заказы»."
                    if is_ajax:
                        return jsonify({"ok": False, "message": msg, "next_url": next_url}), 500
                    flash(msg, "error")
                failed = result.get("failed_postings") or []
                success_msg = (
                    f"Добавлено в поставку «{result.get('shipment_name') or shipment_id}»: "
                    f"{result.get('count', 0)} из {result.get('requested_count', result.get('count', 0))}."
                )
                if failed:
                    success_msg += f" Пропущено из-за ошибок Ozon: {len(failed)}."
                if is_ajax:
                    return jsonify(
                        {
                            "ok": True,
                            "message": success_msg,
                            "next_url": next_url,
                            "failed_count": len(failed),
                            "failed_postings": failed[:10],
                        }
                    )
                flash(success_msg, "success")
            else:
                err = result.get("error") or "Не удалось добавить в поставку"
                if is_ajax:
                    return jsonify({"ok": False, "message": err, "next_url": next_url}), 400
                flash(err, "error")
        except OzonApiError as exc:
            logging.exception("Ошибка Ozon API при добавлении в существующую поставку")
            if is_ajax:
                return jsonify({"ok": False, "message": str(exc), "next_url": next_url}), 500
            flash(str(exc), "error")

        return redirect(next_url)

    @app.get("/orders_wb/shipment/next-name")
    def orders_wb_shipment_next_name():
        return jsonify({"name": service.suggest_next_shipment_name(marketplace="wb")})

    @app.get("/orders_wb/shipment/available")
    def orders_wb_shipment_available():
        shipments = service.get_wb_supplies_available()
        return jsonify({"shipments": shipments})

    @app.post("/orders_wb/shipment/create")
    def orders_wb_shipment_create():
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        ids_raw = request.form.getlist("order_ids")
        next_url = request.form.get("next") or url_for("orders_wb", status="new")
        if not (isinstance(next_url, str) and next_url.startswith("/") and not next_url.startswith("//")):
            next_url = url_for("orders_wb", status="new")
        shipment_name = (request.form.get("shipment_name") or "").strip()
        split_by_warehouse = (request.form.get("split_by_warehouse") or "") == "1"
        order_ids: List[int] = []
        for x in ids_raw:
            try:
                order_ids.append(int(x))
            except (TypeError, ValueError):
                continue
        try:
            result = service.create_wb_shipment_with_orders(
                shipment_name, order_ids, split_by_warehouse=split_by_warehouse
            )
            if result.get("ok"):
                success_msg = (
                    f"WB поставка «{result.get('name')}» создана ({result.get('wb_supply_id')}), "
                    f"добавлено {result.get('count', 0)} из {result.get('requested_count', 0)}."
                )
                if is_ajax:
                    return jsonify({"ok": True, "message": success_msg, "next_url": next_url})
                flash(success_msg, "success")
            else:
                err = result.get("error") or "Не удалось создать WB поставку"
                if is_ajax:
                    return jsonify({"ok": False, "message": err, "next_url": next_url}), 400
                flash(err, "error")
        except WbApiError as exc:
            logging.exception("Ошибка WB API при создании поставки")
            if is_ajax:
                return jsonify({"ok": False, "message": str(exc), "next_url": next_url}), 500
            flash(str(exc), "error")
        return redirect(next_url)

    @app.post("/orders_wb/shipment/add-existing")
    def orders_wb_shipment_add_existing():
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        ids_raw = request.form.getlist("order_ids")
        next_url = request.form.get("next") or url_for("orders_wb", status="new")
        if not (isinstance(next_url, str) and next_url.startswith("/") and not next_url.startswith("//")):
            next_url = url_for("orders_wb", status="new")
        supply_id = (request.form.get("shipment_id") or "").strip()
        if not supply_id:
            if is_ajax:
                return jsonify({"ok": False, "message": "Не выбрана WB поставка", "next_url": next_url}), 400
            flash("Не выбрана WB поставка", "error")
            return redirect(next_url)
        order_ids: List[int] = []
        for x in ids_raw:
            try:
                order_ids.append(int(x))
            except (TypeError, ValueError):
                continue
        try:
            result = service.add_orders_to_existing_wb_supply(supply_id, order_ids)
            if result.get("ok"):
                success_msg = (
                    f"В WB поставку «{result.get('shipment_name') or supply_id}» добавлено "
                    f"{result.get('count', 0)} из {result.get('requested_count', 0)}."
                )
                if is_ajax:
                    return jsonify({"ok": True, "message": success_msg, "next_url": next_url})
                flash(success_msg, "success")
            else:
                err = result.get("error") or "Не удалось добавить в WB поставку"
                if is_ajax:
                    return jsonify({"ok": False, "message": err, "next_url": next_url}), 400
                flash(err, "error")
        except WbApiError as exc:
            logging.exception("Ошибка WB API при добавлении в поставку")
            if is_ajax:
                return jsonify({"ok": False, "message": str(exc), "next_url": next_url}), 500
            flash(str(exc), "error")
        return redirect(next_url)

    @app.get("/shipments")
    def shipments():
        return redirect(url_for("orders"))

    @app.get("/shipments/<int:shipment_id>")
    def shipment_detail(shipment_id: int):
        marketplace = (request.args.get("marketplace") or "ozon").strip().lower()
        if marketplace not in ("ozon", "wb"):
            marketplace = "ozon"
        data = service.get_shipment_detail(shipment_id, marketplace=marketplace)
        shipment = data.get("shipment") or {}
        orders = data.get("orders") or []
        return render_template(
            "shipment_detail.html",
            shipment=shipment,
            orders=orders,
            shipment_is_wb=(marketplace == "wb"),
        )

    @app.get("/shipments/wb/<string:supply_id>")
    def wb_shipment_detail(supply_id: str):
        data = service.get_wb_shipment_detail_by_supply_id(supply_id)
        shipment = data.get("shipment") or {}
        orders = data.get("orders") or []
        return render_template(
            "shipment_detail.html",
            shipment=shipment,
            orders=orders,
            shipment_is_wb=True,
        )

    @app.get("/health")
    def health():
        return jsonify({"status": "ok"})

    @app.cli.command("sync-wb-catalog")
    def sync_wb_catalog_command():
        """Полная выгрузка каталога WB в instance/wb_catalog.db (cron). Параллельный запуск блокируется lock-файлом."""
        import sys

        from services.wb_catalog_service import run_full_wb_catalog_sync

        logging.basicConfig(level=logging.INFO, stream=sys.stderr)

        def _pr(ev):
            print(ev, file=sys.stderr, flush=True)

        stats = run_full_wb_catalog_sync(progress_cb=_pr)
        if not stats.get("skipped"):
            try:
                en = service.enrich_wb_order_items_from_local_catalog()
                stats["order_items_enriched"] = int(en.get("items_updated") or 0)
            except Exception:
                logging.exception("enrich_wb_order_items_from_local_catalog после sync-wb-catalog")
        print(json.dumps(stats, ensure_ascii=False, indent=2))

    @app.cli.command("enrich-wb-items-catalog")
    def enrich_wb_items_catalog_command():
        """Обновить order_items из wb_catalog.db по nmId (без Content API)."""
        import sys

        logging.basicConfig(level=logging.INFO, stream=sys.stderr)
        stats = service.enrich_wb_order_items_from_local_catalog()
        print(json.dumps(stats, ensure_ascii=False, indent=2))

    @app.cli.command("sync-wb-catalog-new")
    def sync_wb_catalog_new_command():
        """Карточки Content API только для заказов «Новые» (WB) → wb_catalog.db + order_items."""
        import sys

        logging.basicConfig(level=logging.INFO, stream=sys.stderr)

        def _pr(ev):
            print(ev, file=sys.stderr, flush=True)

        stats = service.sync_wb_catalog_for_new_orders_only(progress_cb=_pr)
        print(json.dumps(stats, ensure_ascii=False, indent=2))

    @app.cli.command("rebuild-label-pdfs")
    @click.option(
        "--fetch-ozon-sizes/--no-fetch-ozon-sizes",
        default=False,
        help="Для Ozon подтянуть этикетку с API, чтобы подогнать размер страницы (медленно).",
    )
    @click.option(
        "--marketplace",
        "marketplace_filter",
        default=None,
        help="Только заказы этого маркетплейса: ozon или wb (по умолчанию — все).",
    )
    def rebuild_label_pdfs_command(fetch_ozon_sizes: bool, marketplace_filter: Optional[str]):
        """Пересоздать instance/labels/order_*.pdf для всех заказов (после смены шаблона и т.п.)."""
        import sys

        mf = (marketplace_filter or "").strip().lower() or None
        if mf and mf not in ("ozon", "wb"):
            raise click.BadParameter("ожидается ozon или wb", param_hint="--marketplace")

        logging.basicConfig(level=logging.INFO, stream=sys.stderr)
        stats = service.rebuild_all_label_pdfs(
            fetch_ozon_for_size=fetch_ozon_sizes,
            marketplace=mf,
        )
        print(json.dumps(stats, ensure_ascii=False, indent=2), file=sys.stdout)

    return app


app = create_app()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
